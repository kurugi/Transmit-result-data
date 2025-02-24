'''
아래는 이 프로그램에 대한 간단한 요약입니다. 핵심 기능과 동작을 한글로 간결하게 정리했습니다.

---

### 프로그램 요약
- **이름**: 병원결과 → LG전자 전송 프로그램
- **목적**: 병원 검사 결과와 소견 데이터를 LG전자 형식으로 변환.
- **사용 기술**: Python, PyQt6(GUI), openpyxl(Excel 처리).

#### 주요 기능
1. **파일 선택**:
   - 병원 결과 파일(`.xlsx`)과 병원 소견 파일(`.xlsx`)을 선택.
2. **데이터 변환**:
   - 병원 데이터를 LG 템플릿(`LG결과(템플릿).xlsx`)에 맞게 변환.
   - 숫자 변환, 주민번호 자르기, BMI 계산 등 수행.
3. **소견 매핑**:
   - 사원번호(`HO_NO`, `EMP_NO`)와 주민번호(`jumin`, `SSN`)로 매칭.
   - 매칭된 소견 데이터를 LG 파일에 추가.
4. **저장**:
   - 변환된 데이터를 사용자가 지정한 새 Excel 파일에 저장.
5. **진행 표시**:
   - 진행바와 로그로 작업 상태 실시간 확인.

#### 동작 방식
- 실행 → 파일 선택 → "변환 시작" 클릭 → 변환 및 매핑 → 결과 파일 저장.

#### 특징
- **정확성**: 사원번호와 주민번호 결합으로 데이터 매핑 신뢰도 향상.
- **사용 편의성**: 버튼과 진행바로 직관적 조작 가능.
- **오류 관리**: 파일 누락/매칭 실패 시 경고 메시지 제공.



이 코드는 이전 버전과 비교해 `map_matching_rows_to_transformed` 함수가 개선된 버전입니다. 주요 변경 사항은 병원 소견 파일과 LG 파일 간의 데이터 매핑 방식이 `HO_NO`와 `EMP_NO`, 그리고 `jumin`과 `SSN`을 결합한 키를 사용하도록 변경된 점입니다. 아래에서 각 함수를 번호별로 한글로 설명하겠습니다.

---

### 주요 함수 설명

#### 1. `convert_emp_no(value)`
- **설명**: 사원번호를 적절한 형식으로 변환합니다.
- **동작**: 숫자(정수/실수)는 정수로, 그 외는 문자열로 변환.
- **예시**: `123.0` → `123`, `"ABC"` → `"ABC"`.

#### 2. `convert_to_numeric(value)`
- **설명**: 값을 숫자형으로 변환합니다.
- **동작**: 빈 값은 `None`, 가능하면 실수로 변환, 불가능하면 원래 값 유지.
- **예시**: `"12.5"` → `12.5`, `""` → `None`, `"text"` → `"text"`.

#### 3. `truncate_ssn(value)`
- **설명**: 주민등록번호를 앞 8자리로 자릅니다.
- **동작**: 문자열이면 8자리까지, 아니면 원래 값 반환.
- **예시**: `"123456-7890123"` → `"12345678"`.

#### 4. `convert_mdc_date(value)`
- **설명**: 진료일자를 숫자 형식으로 변환합니다.
- **동작**: "-"를 제거하고 정수로 변환, 실패 시 원래 값 유지.
- **예시**: `"2023-01-01"` → `20230101`.

#### 5. `extract_sex_no(ssn_value)`
- **설명**: 주민번호에서 성별 번호를 추출합니다.
- **동작**: 8번째 숫자가 `1,3,5`면 `22`(남성), `2,4,6`이면 `21`(여성), 아니면 `None`.
- **예시**: `"12345678"` → `21` (8이 2라 여성).

#### 6. `calculate_bm06(bm01_value, sex_no)`
- **설명**: BMI 값을 계산합니다.
- **동작**: 키(`bm01_value`, cm)를 미터로 변환 후 `키^2 * 성별값` 계산, 소수점 1자리 반올림.
- **예시**: `170`, `22` → `(1.7 * 1.7 * 22)` ≈ `63.6`.

#### 7. `__init__(self)`
- **설명**: 프로그램 창과 UI를 초기화합니다.
- **동작**: 창 설정, 버튼(파일 선택, 변환, 종료), 진행바, 텍스트창 생성 및 이벤트 연결.
- **예시**: 실행 시 "병원결과 → LG전자 전송 프로그램" 창 표시.

#### 8. `log(self, message)`
- **설명**: 로그 메시지를 텍스트창에 추가합니다.
- **동작**: 작업 진행 상황이나 오류를 기록.
- **예시**: `"병원결과 파일 선택됨: hospital.xlsx"`.

#### 9. `select_hospital_file(self)`
- **설명**: 병원 결과 파일을 선택합니다.
- **동작**: 파일 선택 대화상자 열기, 경로 저장, 로그와 상태 업데이트.
- **예시**: `hospital.xlsx` 선택 → 로그 출력.

#### 10. `select_opinion_file(self)`
- **설명**: 병원 소견 파일을 선택합니다.
- **동작**: 위와 동일, 소견 파일 대상.
- **예시**: `opinion.xlsx` 선택 → 로그 출력.

#### 11. `update_status(self)`
- **설명**: 파일 선택 상태를 표시합니다.
- **동작**: 두 파일 모두 선택 시 "변환 준비 완료" 메시지 출력.
- **예시**: "모든 파일이 선택되었습니다. 변환을 시작할 수 있습니다."

#### 12. `map_and_transfer_data(self)`
- **설명**: 병원 결과 데이터를 LG 템플릿으로 변환합니다.
- **동작**:
  1. 병원 파일과 LG 템플릿 열기.
  2. 열 매핑 생성.
  3. 데이터 변환(숫자, 주민번호, BMI 등) 후 LG 파일에 기록.
  4. 진행바 업데이트 및 저장.
- **예시**: `hospital.xlsx` → `LG결과_변환.xlsx`로 데이터 이동.

#### 13. `map_matching_rows_to_transformed(self, wb_lg, ws_lg, lg_headers)`
- **설명**: 병원 소견 데이터를 LG 파일에 매핑합니다.
- **동작**:
  1. 소견 파일 열기, 필수 열(`HO_NO`, `EMP_NO`, `jumin`, `SSN`) 확인.
  2. `HO_NO|jumin`과 `EMP_NO|SSN`을 결합한 키로 매핑.
  3. 매칭된 데이터만 LG 파일에 기록.
  4. 매칭 안 된 항목 경고 출력.
  5. 진행바 업데이트 및 저장.
- **예시**: `HO_NO=123|jumin=12345678` → `EMP_NO=123|SSN=12345678` 매칭 후 데이터 추가.

#### 14. `run_conversion(self)`
- **설명**: 전체 변환 프로세스를 실행합니다.
- **동작**:
  1. 필수 파일 확인.
  2. 저장 파일 이름 지정.
  3. 병원 데이터 변환(`map_and_transfer_data`) 후 소견 매핑(`map_matching_rows_to_transformed`).
  4. 진행 상황 표시.
- **예시**: "변환 시작" → "변환 완료!".

---

### 변경된 점
- **매핑 개선**: 이전엔 `HO_NO`와 `EMP_NO`만 비교했지만, 이제 `jumin`과 `SSN`도 추가해 더 정확히 매핑.
- **오류 체크 강화**: 필수 열 누락 시 명확한 오류 메시지 출력.
- **디버깅 추가**: 매칭 안 된 항목을 로그로 알려 문제 파악 가능.

---

### 요약
- **도움 함수(1-6)**: 데이터 변환 및 계산 담당.
- **UI 관련(7-11)**: 사용자 인터페이스 설정 및 파일 선택.
- **데이터 처리(12-14)**: 병원 데이터와 소견을 LG 형식으로 변환 및 저장.



'''
import sys
import openpyxl
from openpyxl.styles import Alignment
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QPushButton, 
                            QFileDialog, QTextEdit, QProgressBar, QLabel)
from PyQt6.QtCore import Qt

# 도움 함수 (변경 없음)
def convert_emp_no(value):
    try:
        return int(value) if isinstance(value, (int, float)) and value == int(value) else str(value)
    except ValueError:
        return str(value)

def convert_to_numeric(value):
    try:
        return float(value) if value not in (None, "", " ") else None
    except ValueError:
        return value

def truncate_ssn(value):
    return str(value)[:8] if isinstance(value, str) else value

def convert_mdc_date(value):
    if isinstance(value, str):
        cleaned_value = value.replace("-", "")
        try:
            return int(cleaned_value)
        except ValueError:
            return cleaned_value
    return value

def extract_sex_no(ssn_value):
    if isinstance(ssn_value, str) and len(ssn_value) >= 8:
        sex_digit = ssn_value[7]
        if sex_digit in {"1", "3", "5"}:
            return 22
        elif sex_digit in {"2", "4", "6"}:
            return 21
    return None

def calculate_bm06(bm01_value, sex_no):
    try:
        return round((float(bm01_value) / 100) * (float(bm01_value) / 100) * sex_no, 1)
    except (ValueError, TypeError, ZeroDivisionError):
        return None

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("병원결과 → LG전자 전송 프로그램")
        self.setGeometry(100, 100, 600, 400)

        # 파일 경로
        self.hospital_file = ""
        self.lg_file = "LG결과(템플릿).xlsx"  # LG결과 템플릿 파일 고정
        self.opinion_file = ""
        self.transformed_file = ""  # 사용자가 지정할 예정

        # UI 설정
        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        # 상태 라벨
        self.status_label = QLabel("병원결과와 병원소견 파일을 선택하세요.")
        self.status_label.setAlignment(Qt.AlignmentFlag.AlignCenter)

        # 버튼과 출력창
        self.hospital_btn = QPushButton("병원결과 파일 선택")
        self.opinion_btn = QPushButton("병원소견 파일 선택")
        self.convert_btn = QPushButton("데이터 변환 시작")
        self.exit_btn = QPushButton("종료")  # 종료 버튼

        # 버튼 스타일시트 (기본 배경색 추가, 호버 시 선명한 연두색)
        button_style = """
            QPushButton {
                min-height: 45px;  /* 기본 높이의 1.5배 */
                background-color: #FFFFFF;  /* 기본 배경색: 흰색 */
                border: 1px solid #CCCCCC;  /* 테두리 추가로 구분 */
            }
            QPushButton:hover {
                background-color: #00FF00;  /* 호버 시 선명한 연두색 */
            }
        """
        self.hospital_btn.setStyleSheet(button_style)
        self.opinion_btn.setStyleSheet(button_style)
        self.convert_btn.setStyleSheet(button_style)
        self.exit_btn.setStyleSheet(button_style)

        self.output_text = QTextEdit()
        self.output_text.setReadOnly(True)
        self.progress_bar = QProgressBar()

        self.layout.addWidget(self.status_label)
        self.layout.addWidget(self.hospital_btn)
        self.layout.addWidget(self.opinion_btn)
        self.layout.addWidget(self.convert_btn)
        self.layout.addWidget(self.progress_bar)
        self.layout.addWidget(self.output_text)
        self.layout.addWidget(self.exit_btn)  # 종료 버튼 추가

        # 버튼 연결
        self.hospital_btn.clicked.connect(self.select_hospital_file)
        self.opinion_btn.clicked.connect(self.select_opinion_file)
        self.convert_btn.clicked.connect(self.run_conversion)
        self.exit_btn.clicked.connect(self.close)  # 종료 버튼 동작 연결

        # 열 매핑 및 설정 (기존과 동일)
        self.column_map = {
            "사원번호": "EMP_NO", "주민등록번호": "SSN", "진료일자": "MDC_DATE", "HA001": "BM01",
            "HA002": "BM02", "HA004": "BM04", "FAT": "BM05", "HA007": "BM07", "L90001": "BT01",
            "L70013": "BT02", "HB001": "CV02", "HB002": "CV01", "A8001": "CV03", "E6541": "CV04",
            "HO001": "OE101", "HO002": "OE102", "HO003": "OE103", "HO004": "OE104", "HO011": "OE205",
            "HO005": "OE301", "HO006": "OE302", "OHA01": "AM103", "OHA09": "AM104", "OHA02": "AM105",
            "OHA10": "AM106", "OHA03": "AM107", "OHA11": "AM108", "OHA04": "AM121", "OHA12": "AM122",
            "OHA05": "AM111", "OHA13": "AM112", "OHA06": "AM115", "OHA14": "AM116", "L2012": "CB101",
            "L2014": "CB110", "L07003501": "CB112", "L20220": "CB113", "L20141": "CB104", "L20142": "CB105",
            "L20143": "CB106", "L2015": "CB107", "L2016": "CB108", "L2011": "CB109", "L20191": "CB203",
            "L20193": "CB204", "L20192": "CB205", "L20194": "CB206", "L3012": "DM04", "L3231": "DM07",
            "L3015": "LP01", "L3081": "LP02", "L3082": "LP03", "L3083": "LP04", "L3016": "LF13",
            "LAC10401": "LF14", "A009": "LF15", "L3018": "LF03", "L3019": "LF04", "L3033": "LF05",
            "L3020": "LF06", "L3062": "LF11", "LAC10402": "LF10", "L3051": "LF16", "LAC114": "LF17",
            "LAC158": "LF19", "N20501": "VE105", "N20502": "VE102", "N20511": "VE201", "L170360": "VE106",
            "L170380": "VE107", "L3032": "RF03", "L3013": "RF02", "LAC104031": "RF04", "LAC13701": "EL01",
            "L3031": "EL02", "L3041": "EL03", "L3042": "EL04", "RU103": "TF01", "N20006": "TF06",
            "N20003": "TF03", "LIS11001": "VE301", "L3092": "SY03", "L51821": "CE01", "L51811": "CE02",
            "N206101": "CE03", "N20609": "CE04", "N206151": "CE05", "L150026": "CE06", "L3014": "GT01",
            "L5114": "RA03", "L6103": "UA101", "L6102": "UA102", "L6110": "UA103", "L6104": "UA106",
            "L6107": "UA108", "L6108": "UA301", "L6109": "UA302", "L6105": "UA401", "L61132": "UA201",
            "L61131": "UA202", "L61133": "UA203", "TK531": "SE02", "TK02": "SE01", "TK03": "SE03",
            "RZ901A2": "BD01", "RP201": "RE101", "L9742HPC": "RE200", "RC4011": "GI303", "S1005": "GI201",
            "L5237": "GI203", "C5602": "GI205", "RU401": "US02", "S2322": "US03", "RU902A": "US04",
            "RU505": "US05", "S1008B": "US07", "CTN710": "US08", "N456232": "US09", "RC101": "US10",
            "C5601": "US13", "RP20BBA": "GY03", "RU201": "GY04", "P30001": "GY05", "RC94HL": "RE402",
            "L2013": "CB102", "L3021": "LF12", "L1820": "LF08", "LAC162": "RA01", "TH01": "RA02",
            "L20221": "CB111", "L20190": "CB201", "RC94HC": "RE403", "L6106": "UA402", "L5237": "G1203",
            "RN801": "US15", "N456000": "US16", "N455004": "US17", "CTN711H": "US20", "SE60": "US11",
            "TX0B300": "GY07"
        }
        self.numeric_columns = {
            "MDC_DATE", "BM01", "BM02", "BM04", "BM05", "BM07", "CV02", "CV01", "CV03",
            "OE101", "OE102", "OE103", "OE104", "OE301", "OE302", "AM103", "AM104", "AM105",
            "AM106", "AM107", "AM108", "AM121", "AM122", "AM111", "AM112", "AM115", "AM116",
            "CB101", "CB110", "CB112", "CB113", "CB104", "CB105", "CB106", "CB107", "CB108",
            "CB109", "CB203", "CB204", "CB205", "CB206", "DM04", "DM07", "LP01", "LP02",
            "LP03", "LP04", "LF13", "LF14", "LF15", "LF04", "LF05", "LF06", "LF11", "LF10",
            "LF16", "LF19", "RF03", "RF02", "RF04", "EL01", "EL02", "EL03", "EL04", "TF06",
            "TF03", "VE301", "SY03", "CE01", "CE02", "CE03", "CE04", "CE05", "CE06", "GT01",
            "RA01", "RA02", "RA03", "UA101", "UA102", "CB102", "CB111", "CB201", "LF03",
            "LF12", "LF08", "LF17", "BM06"
        }
        self.right_align_columns = {"CE01", "CE02", "CE03"}
        self.opinion_columns = {
            "A1": "MDC_DECI", "A2": "STATE", "A3": "RECIPE1", "A4": "RECIPE2", "A5": "RECIPE3",
            "A6": "RECIPE4", "A7": "RECIPE5", "A8": "MDC_GRADE1", "A9": "OPIN_CODE1", "A10": "OPIN_DESCRIPT1",
            "A11": "MDC_GRADE2", "A12": "OPIN_CODE2", "A13": "OPIN_DESCRIPT2", "A14": "MDC_GRADE3",
            "A15": "OPIN_CODE3", "A16": "OPIN_DESCRIPT3", "A17": "MDC_GRADE4", "A18": "OPIN_CODE4",
            "A19": "OPIN_DESCRIPT4", "A20": "MDC_GRADE5", "A21": "OPIN_CODE5", "A22": "OPIN_DESCRIPT5"
        }

    def log(self, message):
        self.output_text.append(message)

    def select_hospital_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "병원결과 파일 선택", "", "Excel Files (*.xlsx)")
        if file:
            self.hospital_file = file
            self.log(f"병원결과 파일 선택됨: {file}")
            self.update_status()

    def select_opinion_file(self):
        file, _ = QFileDialog.getOpenFileName(self, "병원소견 파일 선택", "", "Excel Files (*.xlsx)")
        if file:
            self.opinion_file = file
            self.log(f"병원소견 파일 선택됨: {file}")
            self.update_status()

    def update_status(self):
        if self.hospital_file and self.opinion_file:
            self.status_label.setText("모든 파일이 선택되었습니다. 변환을 시작할 수 있습니다.")
        elif self.hospital_file:
            self.status_label.setText("병원결과 파일이 선택되었습니다. 병원소견 파일을 선택하세요.")
        elif self.opinion_file:
            self.status_label.setText("병원소견 파일이 선택되었습니다. 병원결과 파일을 선택하세요.")
        else:
            self.status_label.setText("병원결과와 병원소견 파일을 선택하세요.")

    def map_and_transfer_data(self):
        try:
            wb_hospital = openpyxl.load_workbook(self.hospital_file)
            ws_hospital = wb_hospital.active
        except FileNotFoundError:
            self.log(f"오류: {self.hospital_file} 파일을 찾을 수 없습니다.")
            return None, None, None
        except Exception as e:
            self.log(f"오류: 병원결과 파일 처리 중 문제 발생 - {str(e)}")
            return None, None, None

        try:
            wb_lg = openpyxl.load_workbook(self.lg_file)
            ws_lg = wb_lg.active
        except FileNotFoundError:
            self.log(f"오류: {self.lg_file} 파일을 찾을 수 없습니다.")
            return None, None, None
        except Exception as e:
            self.log(f"오류: LG결과 파일 처리 중 문제 발생 - {str(e)}")
            return None, None, None

        ws_lg.delete_rows(4, ws_lg.max_row)
        wb_lg.save(self.transformed_file)
        self.log(f"{self.transformed_file}의 4행부터 모든 데이터가 삭제되었습니다.")

        hospital_headers = {cell.value: col_idx for col_idx, cell in enumerate(ws_hospital[3], 1) if cell.value}
        lg_headers = {cell.value: col_idx for col_idx, cell in enumerate(ws_lg[3], 1) if cell.value}

        transformed_mapping = {
            hospital_headers[header]: lg_headers[self.column_map[header]]
            for header in self.column_map if header in hospital_headers and self.column_map[header] in lg_headers
        }

        ssn_col_idx = lg_headers.get("SSN")
        bm01_col_idx = lg_headers.get("BM01")
        bm06_col_idx = lg_headers.get("BM06")

        data_rows = list(ws_hospital.iter_rows(min_row=5, values_only=True))
        start_row = 4
        total_rows = len(data_rows)

        self.progress_bar.setMaximum(total_rows)
        for row_idx, row in enumerate(data_rows, start=start_row):
            sex_no = None
            bm01_value = None

            for hospital_col, lg_col in transformed_mapping.items():
                value = row[hospital_col - 1]
                lg_column_name = list(lg_headers.keys())[list(lg_headers.values()).index(lg_col)]

                if self.column_map.get(list(hospital_headers.keys())[list(hospital_headers.values()).index(hospital_col)]) == "EMP_NO":
                    value = convert_emp_no(value)
                if lg_column_name in self.numeric_columns:
                    value = convert_to_numeric(value)
                if lg_column_name == "SSN":
                    value = truncate_ssn(value)
                    sex_no = extract_sex_no(value)
                if lg_column_name == "MDC_DATE":
                    value = convert_mdc_date(value)
                if lg_column_name == "BM01":
                    bm01_value = value

                cell = ws_lg.cell(row=row_idx, column=lg_col, value=value)
                if lg_column_name in self.right_align_columns:
                    cell.alignment = Alignment(horizontal="right")

            if bm06_col_idx and bm01_value and sex_no:
                bm06_value = calculate_bm06(bm01_value, sex_no)
                ws_lg.cell(row=row_idx, column=bm06_col_idx, value=bm06_value)

            self.progress_bar.setValue(row_idx - start_row + 1)

        wb_lg.save(self.transformed_file)
        self.log(f"병원결과 데이터가 변환되어 {self.transformed_file}에 저장되었습니다.")
        return wb_lg, ws_lg, lg_headers

    def map_matching_rows_to_transformed(self, wb_lg, ws_lg, lg_headers):
        try:
            wb_opinion = openpyxl.load_workbook(self.opinion_file)
            ws_opinion = wb_opinion.active
        except FileNotFoundError:
            self.log(f"오류: {self.opinion_file} 파일을 찾을 수 없습니다.")
            return
        except Exception as e:
            self.log(f"오류: 병원소견 파일 처리 중 문제 발생 - {str(e)}")
            return

        opinion_headers = {cell.value: col_idx for col_idx, cell in enumerate(ws_opinion[2], 1) if cell.value}
        lg_headers = {cell.value: col_idx for col_idx, cell in enumerate(ws_lg[3], 1) if cell.value}

        # 필수 열 확인
        ho_no_col = opinion_headers.get("HO_NO")
        emp_no_col = lg_headers.get("EMP_NO")
        jumin_col = opinion_headers.get("jumin")
        ssn_col = lg_headers.get("SSN")

        if not all([ho_no_col, emp_no_col, jumin_col, ssn_col]):
            missing_cols = []
            if not ho_no_col: missing_cols.append("HO_NO")
            if not emp_no_col: missing_cols.append("EMP_NO")
            if not jumin_col: missing_cols.append("jumin")
            if not ssn_col: missing_cols.append("SSN")
            self.log(f"오류: 다음 필수 열이 없습니다: {', '.join(missing_cols)}")
            return

        # HO_NO와 EMP_NO, jumin과 SSN을 결합한 키로 매핑
        ho_no_jumin_dict = {
            f"{str(row[ho_no_col - 1])}|{str(row[jumin_col - 1])}": row 
            for row in ws_opinion.iter_rows(min_row=3, values_only=True) 
            if row[ho_no_col - 1] is not None and row[jumin_col - 1] is not None
        }
        emp_no_ssn_dict = {
            f"{str(row[emp_no_col - 1].value)}|{str(row[ssn_col - 1].value)}": row_idx 
            for row_idx, row in enumerate(ws_lg.iter_rows(min_row=4), 4) 
            if row[emp_no_col - 1].value is not None and row[ssn_col - 1].value is not None
        }

        common_keys = set(ho_no_jumin_dict.keys()) & set(emp_no_ssn_dict.keys())
        self.log(f"공통 데이터 개수 (HO_NO=EMP_NO AND jumin=SSN): {len(common_keys)}")

        # 매칭되지 않은 항목 확인
        unmatched_ho_jumin = set(ho_no_jumin_dict.keys()) - common_keys
        unmatched_emp_ssn = set(emp_no_ssn_dict.keys()) - common_keys
        if unmatched_ho_jumin:
            self.log(f"경고: {len(unmatched_ho_jumin)}개의 HO_NO|jumin 조합이 매칭되지 않음: {list(unmatched_ho_jumin)[:5]}...")
        if unmatched_emp_ssn:
            self.log(f"경고: {len(unmatched_emp_ssn)}개의 EMP_NO|SSN 조합이 매칭되지 않음: {list(unmatched_emp_ssn)[:5]}...")

        opinion_mapping = {
            opinion_headers.get(col): lg_headers.get(self.opinion_columns[col])
            for col in self.opinion_columns
            if col in opinion_headers and self.opinion_columns[col] in lg_headers
        }

        self.progress_bar.setMaximum(len(common_keys))
        for i, key in enumerate(common_keys):
            ho_no_row = ho_no_jumin_dict[key]
            lg_row_idx = emp_no_ssn_dict[key]
            for opinion_col, lg_col in opinion_mapping.items():
                value = ho_no_row[opinion_col - 1]
                ws_lg.cell(row=lg_row_idx, column=lg_col, value=value)
            self.progress_bar.setValue(i + 1)

        wb_lg.save(self.transformed_file)
        self.log(f"소견 데이터가 매핑되어 {self.transformed_file}에 저장되었습니다.")

    def run_conversion(self):
        if not all([self.hospital_file, self.lg_file, self.opinion_file]):
            self.log("변환을 시작하기 전에 모든 필수 파일을 선택해주세요.")
            self.status_label.setText("파일이 모두 선택되지 않았습니다.")
            return

        # 사용자가 저장 파일 이름을 지정하도록 대화 상자 열기
        self.transformed_file, _ = QFileDialog.getSaveFileName(self, "저장할 파일 이름 지정", "LG결과_변환.xlsx", "Excel Files (*.xlsx)")
        if not self.transformed_file:
            self.log("저장 파일 이름이 지정되지 않았습니다. 변환을 취소합니다.")
            self.status_label.setText("저장 파일 이름이 지정되지 않았습니다.")
            return

        self.log("변환 시작...")
        self.status_label.setText("데이터 변환 중...")
        self.progress_bar.setValue(0)
        wb_lg, ws_lg, lg_headers = self.map_and_transfer_data()
        if wb_lg and ws_lg and lg_headers:
            self.map_matching_rows_to_transformed(wb_lg, ws_lg, lg_headers)
        self.log("변환 완료!")
        self.status_label.setText("변환 완료! 결과를 확인하세요.")

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


    # pyinstaller --onefile --noconsole LG결과_PyQt6.py