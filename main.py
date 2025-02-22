import openpyxl
from openpyxl.styles import Alignment

# 도움 함수 정의
def convert_emp_no(value):
    """EMP_NO가 숫자형이면 숫자로, 문자형이면 그대로 반환"""
    try:
        return int(value) if isinstance(value, (int, float)) and value == int(value) else str(value)
    except ValueError:
        return str(value)

def convert_to_numeric(value):
    """숫자형 데이터 변환 (변환 가능하면 float, 불가능하면 그대로)"""
    try:
        return float(value) if value not in (None, "", " ") else None
    except ValueError:
        return value

def truncate_ssn(value):
    """주민등록번호(SSN) 왼쪽 8자리만 출력"""
    return str(value)[:8] if isinstance(value, str) else value

def convert_mdc_date(value):
    """MDC_DATE에서 '-' 제거 후 숫자로 변환"""
    if isinstance(value, str):
        cleaned_value = value.replace("-", "")
        try:
            return int(cleaned_value)
        except ValueError:
            return cleaned_value
    return value

def extract_sex_no(ssn_value):
    """SSN 값에서 왼쪽에서 8번째 문자를 추출하여 sex_no 변수에 저장"""
    if isinstance(ssn_value, str) and len(ssn_value) >= 8:
        sex_digit = ssn_value[7]
        if sex_digit in {"1", "3", "5"}:
            return 22
        elif sex_digit in {"2", "4", "6"}:
            return 21
    return None

def calculate_bm06(bm01_value, sex_no):
    """BM06 값을 계산"""
    try:
        return round((float(bm01_value) / 100) * (float(bm01_value) / 100) * sex_no, 1)
    except (ValueError, TypeError, ZeroDivisionError):
        return None

# 병원결과 -> LG결과_변환 함수
def map_and_transfer_data(hospital_file, lg_file, output_file, column_map, numeric_columns, right_align_columns):
    try:
        wb_hospital = openpyxl.load_workbook(hospital_file)
        ws_hospital = wb_hospital.active
    except FileNotFoundError:
        print(f"에러: {hospital_file} 파일을 찾을 수 없습니다.")
        return

    try:
        wb_lg = openpyxl.load_workbook(lg_file)
        ws_lg = wb_lg.active
    except FileNotFoundError:
        print(f"에러: {lg_file} 파일을 찾을 수 없습니다.")
        return

    # 4행부터 데이터 삭제
    ws_lg.delete_rows(4, ws_lg.max_row)
    wb_lg.save(output_file)
    print(f"4행부터 모든 데이터가 {output_file}에서 삭제되었습니다.")

    # 병원결과 헤더 (3번째 행)
    hospital_headers = {cell.value: col_idx for col_idx, cell in enumerate(ws_hospital[3], 1) if cell.value}
    # LG결과 헤더 (3번째 행)
    lg_headers = {cell.value: col_idx for col_idx, cell in enumerate(ws_lg[3], 1) if cell.value}

    # 매핑 적용
    transformed_mapping = {
        hospital_headers[header]: lg_headers[column_map[header]]
        for header in column_map if header in hospital_headers and column_map[header] in lg_headers
    }

    ssn_col_idx = lg_headers.get("SSN", None)
    bm01_col_idx = lg_headers.get("BM01", None)
    bm06_col_idx = lg_headers.get("BM06", None)

    data_rows = list(ws_hospital.iter_rows(min_row=5, values_only=True))
    start_row = 4

    for row_idx, row in enumerate(data_rows, start=start_row):
        sex_no = None
        bm01_value = None

        for hospital_col, lg_col in transformed_mapping.items():
            value = row[hospital_col - 1]
            lg_column_name = list(lg_headers.keys())[list(lg_headers.values()).index(lg_col)]

            if column_map.get(list(hospital_headers.keys())[list(hospital_headers.values()).index(hospital_col)]) == "EMP_NO":
                value = convert_emp_no(value)
            if lg_column_name in numeric_columns:
                value = convert_to_numeric(value)
            if lg_column_name == "SSN":
                value = truncate_ssn(value)
                sex_no = extract_sex_no(value)
            if lg_column_name == "MDC_DATE":
                value = convert_mdc_date(value)
            if lg_column_name == "BM01":
                bm01_value = value

            cell = ws_lg.cell(row=row_idx, column=lg_col, value=value)
            if lg_column_name in right_align_columns:
                cell.alignment = Alignment(horizontal="right")

        if bm06_col_idx and bm01_value and sex_no:
            bm06_value = calculate_bm06(bm01_value, sex_no)
            ws_lg.cell(row=row_idx, column=bm06_col_idx, value=bm06_value)

    wb_lg.save(output_file)
    print(f"변환된 데이터가 {output_file}에 저장되었습니다.")
    return wb_lg, ws_lg, lg_headers  # 매핑에 사용하기 위해 반환

# 병원소견 -> LG결과_변환 매핑 함수
def map_matching_rows_to_transformed(opinion_file, wb_lg, ws_lg, lg_headers, opinion_columns):
    try:
        wb_opinion = openpyxl.load_workbook(opinion_file)
        ws_opinion = wb_opinion.active
    except FileNotFoundError:
        print(f"에러: {opinion_file} 파일을 찾을 수 없습니다.")
        return

    opinion_headers = {cell.value: col_idx for col_idx, cell in enumerate(ws_opinion[2], 1) if cell.value}
    ho_no_col = opinion_headers.get("HO_NO")
    if not ho_no_col:
        print(f"'HO_NO'가 {opinion_file}의 2행에 존재하지 않습니다.")
        return
    else:
        print(f"'HO_NO'가 {opinion_file}의 2행 {ho_no_col}번째 열에 존재합니다.")

    emp_no_col = lg_headers.get("EMP_NO")
    if not emp_no_col:
        print(f"'EMP_NO'가 LG결과_변환의 3행에 존재하지 않습니다.")
        return
    else:
        print(f"'EMP_NO'가 LG결과_변환의 3행 {emp_no_col}번째 열에 존재합니다.")

    # opinion_columns 매핑 준비
    opinion_mapping = {
        opinion_headers.get(col): lg_headers.get(opinion_columns[col])
        for col in opinion_columns
        if col in opinion_headers and opinion_columns[col] in lg_headers
    }
    if not opinion_mapping:
        print("매핑할 열이 없습니다. opinion_columns의 키와 값이 두 파일에 존재하는지 확인하세요.")
        return

    # 병원소견 "HO_NO" 데이터 추출
    ho_no_dict = {}
    for row_idx, row in enumerate(ws_opinion.iter_rows(min_row=3, values_only=True), start=3):
        ho_no_value = row[ho_no_col - 1]
        if ho_no_value is not None:
            ho_no_dict[str(ho_no_value)] = row

    # LG결과_변환 "EMP_NO" 데이터 추출
    emp_no_dict = {}
    for row_idx, row in enumerate(ws_lg.iter_rows(min_row=4), start=4):
        emp_no_value = row[emp_no_col - 1].value
        if emp_no_value is not None:
            emp_no_dict[str(emp_no_value)] = row_idx

    # 공통 데이터 매핑
    print("\n=== HO_NO와 EMP_NO가 같은 행의 데이터 매핑 결과 ===")
    common_keys = set(ho_no_dict.keys()) & set(emp_no_dict.keys())

    if common_keys:
        for key in sorted(common_keys):
            ho_no_row = ho_no_dict[key]
            lg_row_idx = emp_no_dict[key]

            print(f"\nHO_NO/EMP_NO 값: {key}")
            print(f"'{opinion_file}' - 행 데이터:")
            print("  ", [val for val in ho_no_row])
            print(f"'LG결과_변환' - 행 {lg_row_idx}에 업데이트:")

            for opinion_col, lg_col in opinion_mapping.items():
                value = ho_no_row[opinion_col - 1]
                ws_lg.cell(row=lg_row_idx, column=lg_col, value=value)
                opinion_col_name = list(opinion_headers.keys())[list(opinion_headers.values()).index(opinion_col)]
                lg_col_name = list(lg_headers.keys())[list(lg_headers.values()).index(lg_col)]
                print(f"  {opinion_col_name} -> {lg_col_name}: {value}")

        wb_lg.save(transformed_file)
        print(f"\n'{transformed_file}' 파일이 업데이트되었습니다.")
    else:
        print("HO_NO와 EMP_NO에 공통 데이터 값이 없습니다.")

    print(f"\n=== 요약 ===")
    print(f"'HO_NO' 데이터 개수: {len(ho_no_dict)}")
    print(f"'EMP_NO' 데이터 개수: {len(emp_no_dict)}")
    print(f"공통 데이터 개수: {len(common_keys)}")

# 설정값
right_align_columns = {"CE01", "CE02", "CE03"}
numeric_columns = {
    "MDC_DATE", "BM01", "BM02", "BM04", "BM05", "BM07", "CV02", "CV01", "CV03",
    "OE101", "OE102", "OE103", "OE104", "OE301", "OE302", "AM103", "AM104", "AM105",
    "AM106", "AM107", "AM108", "AM121", "AM122", "AM111", "AM112", "AM115", "AM116",
    "CB101", "CB110", "CB112", "CB113", "CB104", "CB105", "CB106", "CB107", "CB108",
    "CB109", "CB203", "CB204", "CB205", "CB206", "DM04", "DM07", "LP01", "LP02",
    "LP03", "LP04", "LF13", "LF14", "LF15", "LF04", "LF05", "LF06", "LF11", "LF10",
    "LF16", "LF19", "RF03", "RF02", "RF04", "EL01", "EL02", "EL03", "EL04", "TF06",
    "TF03", "VE301", "SY03", "CE01", "CE02", "CE03", "CE04", "CE05", "CE06", "GT01",
    "RA01", "RA02", "RA03", "UA101", "UA102", "CB102", "CB111", "CB201", "LF03",
    "LF12", "LF08", "LF17", "BM06"
}
column_map = {
    "사원번호": "EMP_NO", "주민등록번호": "SSN", "진료일자": "MDC_DATE", "HA001": "BM01",
    "HA002": "BM02", "HA004": "BM04", "FAT": "BM05", "HA007": "BM07", "L90001": "BT01",
    "L70013": "BT02", "HB001": "CV02", "HB002": "CV01", "A8001": "CV03", "E6541": "CV04",
    "HO001": "OE101", "HO002": "OE102", "HO003": "OE103", "HO004": "OE104", "HO011": "OE205",
    "HO005": "OE301", "HO006": "OE302", "OHA01": "AM103", "OHA09": "AM104", "OHA02": "AM105",
    "OHA10": "AM106", "OHA03": "AM107", "OHA11": "AM108", "OHA04": "AM121", "OHA12": "AM122",
    "OHA05": "AM111", "OHA13": "AM112", "OHA06": "AM115", "OHA14": "AM116", "L2012": "CB101",
    "L2014": "CB110", "L07003501": "CB112", "L20220": "CB113", "L20141": "CB104", "L20142": "CB105",
    "L20143": "CB106", "L2015": "CB107", "L2016": "CB108", "L2011": "CB109", "L20191": "CB203",
    "L20193": "CB204", "L20192": "CB205", "L20194": "CB206", "L3012": "DM04", "L3231": "DM07",
    "L3015": "LP01", "L3081": "LP02", "L3082": "LP03", "L3083": "LP04", "L3016": "LF13",
    "LAC10401": "LF14", "A009": "LF15", "L3018": "LF03", "L3019": "LF04", "L3033": "LF05",
    "L3020": "LF06", "L3062": "LF11", "LAC10402": "LF10", "L3051": "LF16", "LAC114": "LF17",
    "LAC158": "LF19", "N20501": "VE105", "N20502": "VE102", "N20511": "VE201", "L170360": "VE106",
    "L170380": "VE107", "L3032": "RF03", "L3013": "RF02", "LAC104031": "RF04", "LAC13701": "EL01",
    "L3031": "EL02", "L3041": "EL03", "L3042": "EL04", "RU103": "TF01", "N20006": "TF06",
    "N20003": "TF03", "LIS11001": "VE301", "L3092": "SY03", "L51821": "CE01", "L51811": "CE02",
    "N206101": "CE03", "N20609": "CE04", "N206151": "CE05", "L150026": "CE06", "L3014": "GT01",
    "L5114": "RA03", "L6103": "UA101", "L6102": "UA102", "L6110": "UA103", "L6104": "UA106",
    "L6107": "UA108", "L6108": "UA301", "L6109": "UA302", "L6105": "UA401", "L61132": "UA201",
    "L61131": "UA202", "L61133": "UA203", "TK531": "SE02", "TK02": "SE01", "TK03": "SE03",
    "RZ901A2": "BD01", "RP201": "RE101", "L9742HPC": "RE200", "RC4011": "GI303", "S1005": "GI201",
    "L5237": "GI203", "C5602": "GI205", "RU401": "US02", "S2322": "US03", "RU902A": "US04",
    "RU505": "US05", "S1008B": "US07", "CTN710": "US08", "N456232": "US09", "RC101": "US10",
    "C5601": "US13", "RP20BBA": "GY03", "RU201": "GY04", "P30001": "GY05", "RC94HL": "RE402",
    "L2013": "CB102", "L3021": "LF12", "L1820": "LF08", "LAC162": "RA01", "TH01": "RA02",
    "L20221": "CB111", "L20190": "CB201", "RC94HC": "RE403", "L6106": "UA402", "L5237": "G1203",
    "RN801": "US15", "N456000": "US16", "N455004": "US17", "CTN711H": "US20", "SE60": "US11",
    "TX0B300": "GY07"
}
opinion_columns = {
    "A1": "MDC_DECI", "A2": "STATE", "A3": "RECIPE1", "A4": "RECIPE2", "A5": "RECIPE3",
    "A6": "RECIPE4", "A7": "RECIPE5", "A8": "MDC_GRADE1", "A9": "OPIN_CODE1", "A10": "OPIN_DESCRIPT1",
    "A11": "MDC_GRADE2", "A12": "OPIN_CODE2", "A13": "OPIN_DESCRIPT2", "A14": "MDC_GRADE3",
    "A15": "OPIN_CODE3", "A16": "OPIN_DESCRIPT3", "A17": "MDC_GRADE4", "A18": "OPIN_CODE4",
    "A19": "OPIN_DESCRIPT4", "A20": "MDC_GRADE5", "A21": "OPIN_CODE5", "A22": "OPIN_DESCRIPT5"
}

# 실행
hospital_file = "병원결과.xlsx"
lg_file = "LG결과.xlsx"
opinion_file = "병원소견.xlsx"
transformed_file = "LG결과_변환.xlsx"

# 1단계: 병원결과 데이터를 LG결과_변환으로 변환
wb_lg, ws_lg, lg_headers = map_and_transfer_data(hospital_file, lg_file, transformed_file, column_map, numeric_columns, right_align_columns)

# 2단계: 병원소견 데이터를 매핑
if wb_lg and ws_lg and lg_headers:
    map_matching_rows_to_transformed(opinion_file, wb_lg, ws_lg, lg_headers, opinion_columns)